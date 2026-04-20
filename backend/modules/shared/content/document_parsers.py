"""共享文档解析器，将上传文件提取为纯文本，供知识库和文档导入复用。"""

from __future__ import annotations

import io

from core.exceptions import ValidationError


def parse_uploaded_text(*, file_content: bytes, filename: str) -> str:
    """按文件后缀解析上传文本，输出标准化纯文本。"""
    filename_lower = filename.lower()
    if filename_lower.endswith(".txt"):
        return _parse_txt(file_content)
    if filename_lower.endswith(".md"):
        return _parse_md(file_content)
    if filename_lower.endswith(".docx"):
        return _parse_docx(file_content)
    if filename_lower.endswith(".pdf"):
        return _parse_pdf(file_content)
    if filename_lower.endswith(".xlsx"):
        return _parse_xlsx(file_content)
    raise ValidationError(
        message="文件格式不支持",
        field_errors={"file": "仅支持 .txt、.md、.docx、.pdf、.xlsx 格式"},
    )


def _parse_txt(file_content: bytes) -> str:
    """解析 UTF-8 文本文件。"""
    try:
        content = file_content.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValidationError(
            message="文件编码错误", field_errors={"file": "文件必须是 UTF-8 编码"}
        ) from exc
    return _ensure_non_empty(content)


def _parse_md(file_content: bytes) -> str:
    """解析 Markdown 文件，保留原始 Markdown 文本供后续 HTML 转换。"""
    try:
        content = file_content.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValidationError(
            message="文件编码错误",
            field_errors={"file": "Markdown 文件必须是 UTF-8 编码"},
        ) from exc
    return _ensure_non_empty(content)


def _parse_docx(file_content: bytes) -> str:
    """解析 Word 文档为段落文本。"""
    try:
        from docx import Document
    except ImportError as exc:
        raise ValidationError(
            message="服务器缺少依赖",
            field_errors={"file": "服务器未安装 python-docx 库"},
        ) from exc
    try:
        document = Document(io.BytesIO(file_content))
    except Exception as exc:
        raise ValidationError(
            message="Word 文件解析失败",
            field_errors={"file": "请确认上传的是有效的 .docx 文件"},
        ) from exc
    content = "\n".join(paragraph.text for paragraph in document.paragraphs)
    return _ensure_non_empty(content)


def _parse_pdf(file_content: bytes) -> str:
    """使用 pypdf 解析 PDF 文本。"""
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValidationError(
            message="服务器缺少依赖",
            field_errors={"file": "服务器未安装 pypdf 库"},
        ) from exc
    try:
        reader = PdfReader(io.BytesIO(file_content))
        content = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
        raise ValidationError(
            message="PDF 文件解析失败",
            field_errors={"file": "请确认上传的是有效的 .pdf 文件"},
        ) from exc
    return _ensure_non_empty(content)


def _parse_xlsx(file_content: bytes) -> str:
    """使用 openpyxl 解析 Excel 文本。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            message="服务器缺少依赖",
            field_errors={"file": "服务器未安装 openpyxl 库"},
        ) from exc
    try:
        workbook = load_workbook(
            io.BytesIO(file_content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ValidationError(
            message="Excel 文件解析失败",
            field_errors={"file": "请确认上传的是有效的 .xlsx 文件"},
        ) from exc
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[{sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                lines.append("\t".join(values))
    return _ensure_non_empty("\n".join(lines))


def _ensure_non_empty(content: str) -> str:
    """对解析结果做统一非空校验并规范空白。"""
    normalized = content.strip()
    if not normalized:
        raise ValidationError(
            message="文件内容为空", field_errors={"file": "上传的文件没有有效内容"}
        )
    return normalized
